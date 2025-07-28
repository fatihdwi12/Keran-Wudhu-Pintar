# VERSI PYTHON 3.9.2


import os



try:

    import cv2 

    import time

    from ultralytics import YOLO

    import requests

    import serial

    import serial.tools.list_ports

    from datetime import date

    from openpyxl import Workbook, load_workbook

    from openpyxl.utils import get_column_letter

    from datetime import datetime

    import json

    import os





except ImportError:

    print("Library tidak ditemukan, menginstal sekarang...")

    os.system("pip install openpyxl==3.1.5 numpy==2.0.0 matplotlib==3.9.2 tqdm==4.67.0 opencv-python==4.10.0.84 ultralytics==8.3.32 requests==2.32.3 pyserial==3.5 --break-system-packages")





    import cv2

    import time

    from ultralytics import YOLO

    import requests

    import serial

    import serial.tools.list_ports

    from datetime import date

    from openpyxl import Workbook, load_workbook

    from openpyxl.utils import get_column_letter

    from datetime import datetime

    import os





######################## DEKLARASI ########################



# rasberry gak bisa cari port otomatis

port_serial = "COM7"
url = "https://localhost.go-web.my.id/2025-fatih-deteksi-gerakkan-wudhu/api/api.php"



gerakan_wudhu = {

    'berniat': 0,

    'membasuh_wajah': 0,

    'membasuh_lengan': 0,

    # 'membasuh_lengan_atas': 0,          # tidak wajib, 

    'membasuh_sebagian_kepala': 0,

    'membasuh_kaki': 0,

    # 'membasuh_betis': 0                 # tidak wajib, 

}





#kode serial untuk mp3

"""

1. berniat
2. membasuh wajah
3. membasuh lengan
4. membasuh sebagian kepala
5. membasuh kaki
6. sistem aktif
7. silakan mendekat
8. gerakan salah
9. wudhu selesai
10. urutan salah
11.  membasuh lengan atas
12. membasuh betis

"""







# Mendapatkan folder saat ini

current_directory = os.path.dirname(os.path.abspath(__file__))



print("Tanggal :", str(date.today()))

print("Folder saat ini:", current_directory)



config_path = os.path.join(current_directory, 'config.json')
path_gambar = os.path.join(current_directory, "gerakan.jpg")



# Default config

default_config = {

    "index_kamera": 0,

    "baudrate": 115200,

    "model": "best.pt"

}



# Cek dan buat file config jika belum ada

if not os.path.exists(config_path):

    with open(config_path, 'w') as file:

        json.dump(default_config, file, indent=4)



# Baca config

with open(config_path, 'r') as file:

    config = json.load(file)



# Inisialisasi kamera berdasarkan config

print("LOADING KAMERA ... ")

cam = cv2.VideoCapture(config["index_kamera"])

#cam.set(cv2.CAP_PROP_BUFFERSIZE, 4)



if not cam.isOpened():

    print("Gagal membuka kamera.")

    exit()



# Resolusi & parameter lain

resolusi = (640, 480)

frame_ready, frame = cam.read()

if not frame_ready or frame is None:

    print("Gagal membaca video.")

    exit()

frame = cv2.resize(frame, resolusi, interpolation=cv2.INTER_AREA)

height, width, _ = frame.shape



# Warna BGR yang dikenali

warna_bgr = [

    ("merah", (0, 0, 255)),

    ("kuning", (0, 255, 255)),

    ("hijau", (0, 255, 0)),

    ("biru", (255, 0, 0)),

    ("ungu", (128, 0, 128)),

    ("putih", (255, 255, 255)),

    ("orange", (0, 165, 255)),

    ("cyan", (255, 255, 0)),

    ("magenta", (255, 0, 255))

]



# Inisialisasi model

model = YOLO(os.path.join(current_directory, config["model"]))

timer_deteksi = 5



# Serial Port

ports = serial.tools.list_ports.comports()

baud_rate = config["baudrate"]



# List dictionary untuk objek per frame

object_per_frame = []

jumlah_frame_kosong = 0





# dummy

timer_audio1 = 0

timer_audio2 = 0

timer_audio3 = 0 #untuk urutan salah

timer_audio4 = 0 #untuk ggerakan salah

last_index_gerakan = 0



interval_audio_silakan = 10

interval_audio_gagal = 10

interval_audio_selesai = 10










def proses_kirim_serial(pesan):

    global ser

    if pesan == "inisiasi":

        try:            

            

            print("Mencoba port ", port_serial)

            ser = serial.Serial(port_serial, baud_rate,

                                        timeout=5, write_timeout=5)

            print(f"Terhubung ke {port_serial} dengan baud rate {baud_rate}")

            time.sleep(10)

                    

        except:

            print("Gagal Terhubung ", port_serial)

            time.sleep(0.1)



            if len(ports) == 0:

                print("TIDAK ADA PORT SERIAL ")

            # Menunggu beberapa saat agar koneksi stabil

            time.sleep(2)

       

    else:

        for t in range(1):

            ser.write(pesan.encode())
            ser.flush()

            print("Mengirim pesan serial : ", pesan.encode())

            time.sleep(0.5)





            

def proses_terima_Serial():

    if ser.in_waiting > 0:

        data = ser.readline().decode('utf-8').strip()

        print("Data Serial diterima:", data)

        return data







def iou(boxA, boxB):

    # Menghitung Intersection over Union (IoU) 

    # untuk penghitung timpang tindih

    xA = max(boxA[0], boxB[0])

    yA = max(boxA[1], boxB[1])

    xB = min(boxA[2], boxB[2])

    yB = min(boxA[3], boxB[3])

    

    interArea = max(0, xB - xA) * max(0, yB - yA)

    boxAArea = (boxA[2] - boxA[0]) * (boxA[3] - boxA[1])

    boxBArea = (boxB[2] - boxB[0]) * (boxB[3] - boxB[1])

    

    iou_value = interArea / float(boxAArea + boxBArea - interArea + 1e-6)

    return iou_value



def non_max_suppression(objs, iou_threshold=0.5):

    # untuk penghitung timpang tindih

    if not objs:

        return []

    

    objs = sorted(objs, key=lambda x: x["confidence"], reverse=True)

    hasil = []

    

    while objs:

        objek_terpilih = objs.pop(0)

        hasil.append(objek_terpilih)

        objs = [

            o for o in objs

            if iou(objek_terpilih["box"], o["box"]) < iou_threshold

        ]

    return hasil



def proses_deteksi(res, frame):

    global object_per_frame



    anti_timpang_tindih = 1 # untuk mengabaikan objek yang terdeteksi tumpang tindih



    frame = cv2.resize(frame, resolusi, interpolation=cv2.INTER_AREA)

    frame_shot = frame.copy()



    results = model(frame)

    object_per_frame = []

    

    for r in results:

        object_per_frame = []

        boxes = (r.boxes.xyxy).tolist()

        for box, conf, cls in zip(boxes, r.boxes.conf, r.boxes.cls):

            x1, y1, x2, y2 = map(int, box)

            label = model.names[int(cls)]

            koordinat = (x1 + (x2 - x1) // 2, y1 + (y2 - y1) // 2)

            

            object_per_frame.append({

                "label": label,

                "koordinat": koordinat,

                "box": [x1, y1, x2, y2],

                "confidence": float(conf),

                "cls": int(cls)

            })



        if anti_timpang_tindih:

            object_per_frame = non_max_suppression(object_per_frame, iou_threshold=0.5)



        # Gambar ke frame

        for obj in object_per_frame:

            x1, y1, x2, y2 = obj["box"]

            label = obj["label"]

            cls = obj["cls"]

            conf = obj["confidence"]



            cv2.putText(frame, f"{label} {str(round(conf*100,2))} %", (x1, obj["koordinat"][1]),

                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 55), 2)

            cv2.rectangle(frame, (x1, y1), (x2, y2), warna_bgr[cls][1], 2)



    return frame



def proses_kirim_gambar(url,path_gambar,deskripsi,status):

    # untuk kirim gambar ke server

    print(" .................. KIRIM GAMBAR....................")
    print(url,path_gambar,deskripsi,status)

    

    try:

        params = {}

        params = {"deskripsi": deskripsi,

                  "status": status} 

        

        with open(path_gambar, 'rb') as j:

            files = {'foto': j}

            response = requests.post(

                url=url, data=params, files=files)

            response.raise_for_status()  # Menambahkan pengecekan status respon



        print(response)
        print(response.text)

    except requests.exceptions.RequestException as e:

        print(f"Gagal mengirim gambar: {e}")





def proses_kirim_data(url):

    # untuk kirim gambar ke server

    print(" .................. KIRIM ....................")

    print("URL : " , url)

    try:



        response = requests.get(

                url=url)

        print(response)

        print(response.text)

    except requests.exceptions.RequestException as e:

        print(f"Gagal mengirim data: {e}")

    

    time.sleep(3)







# Penyimpanan waktu tampil untuk tiap teks

active_texts = {}



def add_timed_text(text="normal", coord=(10, 10), duration_sec=5):

    global active_texts

    now = time.time()

    active_texts[text] = (now, duration_sec, coord)



def draw_active_texts(frame):

    global active_texts

    now = time.time()

    output_frame = frame.copy()



    for t, (start_time, dur, (x, y)) in list(active_texts.items()):

        if now - start_time <= dur:

            cv2.putText(output_frame, t, (x, y), 

                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 55), 2)

        else:

            del active_texts[t]



    return output_frame



def is_berurutan(lst):

    # Harus dari kiri ke kanan, tidak boleh lompat dari 0 ke 1 di posisi lebih tinggi

    for i in range(1, len(lst)):

        if lst[i] == 1 and lst[i-1] == 0:

            # Jika posisi i bernilai 1, tapi posisi sebelumnya 0, maka pastikan semua sebelum i bernilai 1

            if not all(lst[:i]):

                return False

    return True









    





def reset_list(lst):

    return [0] * len(lst)





def proses_logika():

    global tidak_urut, timer_tidak_urut, gerakan_selesai, timer_gerakan_selesai

    global jumlah_frame_kosong, gerakan_wudhu, timer_audio1, timer_audio2, timer_audio3, timer_audio4

    global last_index_gerakan, interval_audio_silakan, interval_audio_gagal, interval_audio_selesai, object_per_frame



    # Inisialisasi awal

    proses_kirim_serial("6,")

    time.sleep(3)

    tidak_urut = 0

    timer_tidak_urut = 0

    gerakan_selesai = 0

    timer_gerakan_selesai = 0



    while True:

        res, frame = cam.read()

        if not res:

            continue



        # Preprocessing

        frame = cv2.convertScaleAbs(frame, alpha=0.8, beta=1)

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        frame = cv2.merge([gray, gray, gray])



        





        if tidak_urut == 0 and gerakan_selesai == 0:

            # Deteksi objek

            frame = proses_deteksi(res, frame)



            # Update status gerakan
            nama_gerakan = ""

            for data in object_per_frame:
                print("data : ",data)
                nama_gerakan = data["label"]
                if data["label"] == "membasuh_lengan_atas" or data["label"] == "membasuh_betis" :
                    if timer_audio4 == 0:
                        proses_kirim_serial("8,") #gerakan salah 
                        
                        timer_audio4 = 1
                        cv2.imwrite(os.path.join(current_directory,"gerakan.jpg"), frame)
                        proses_kirim_gambar(url,path_gambar,nama_gerakan,"gerakkan salah")
                        time.sleep(2)
                
                else :
                    if data["label"] in gerakan_wudhu:
                        gerakan_wudhu[data["label"]] = 1
                        timer_audio4 = 0
                        index_gerakan = list(gerakan_wudhu.keys()).index(data["label"]) + 1


                
            dummy_list = [int(status) for status in gerakan_wudhu.values()]


            if len(object_per_frame) < 1 :

                index_gerakan = -1

            # else :

            #     index_gerakan = sum(dummy_list)





            # Tampilkan status

            posisi_y = 50

            for gerakan, status in gerakan_wudhu.items():

                warna = warna_bgr[2][1] if status == 1 else warna_bgr[1][1]

                posisi_y += 30

                cv2.putText(frame, f"{gerakan}: {status}", (20, posisi_y),

                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, warna, 2)



            # Audio saat terjadi perubahan status

            if last_index_gerakan != index_gerakan:

                timer_audio2 = 0

                last_index_gerakan = index_gerakan



            if timer_audio2 == 0:

                if index_gerakan >= 0 :

                    proses_kirim_serial(f"{index_gerakan},")

                    # Menyimpan gambar ke file baru
                    cv2.imwrite(os.path.join(current_directory,"gerakan.jpg"), frame)
                    proses_kirim_gambar(url,path_gambar,nama_gerakan,"gerakkan benar")


                    time.sleep(1)

                    timer_audio2 = 1

                    timer_audio3 = 0



            # ðŸ”§ CEK URUTAN YANG SALAH

            if not is_berurutan(dummy_list.copy()):  # pakai .copy() agar tidak ubah list asli

                print("TIDAK BERURUTAN")

                if timer_audio3 == 0:

                    proses_kirim_serial("10,")

                    tidak_urut = 1

                    timer_audio3 = 1

                    timer_tidak_urut = time.time()



            # âœ… SEMUA GERAKAN SELESAI

            if all(dummy_list):

                print("SEMUA GERAKAN WUDHU SELESAI")

                gerakan_selesai = 1

                timer_audio3 = 0

                proses_kirim_serial("9,")

                timer_gerakan_selesai = time.time()



        # RESET saat tidak_urut atau selesai

        if tidak_urut or gerakan_selesai:





            now = time.time()

            timer = (now - timer_tidak_urut) if tidak_urut else (now - timer_gerakan_selesai)

            warna = warna_bgr[0][1] if tidak_urut else warna_bgr[2][1]

            cv2.putText(frame, f"Mulai awal ... {round(timer, 2)}", (20, 40),

                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, warna, 2)

            timer_audio3 = 0

            object_per_frame = []

            index_gerakan = 0

            last_index_gerakan = 0



            if (tidak_urut and timer > interval_audio_gagal) or (gerakan_selesai and timer > interval_audio_selesai):

                tidak_urut = 0

                gerakan_selesai = 0

                proses_kirim_serial("6,")

                for gerakan in gerakan_wudhu:

                    gerakan_wudhu[gerakan] = 0



        # Tampilkan frame

        cv2.imshow("Object Detection", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):

            break



    cam.release()

    cv2.destroyAllWindows()









    

if __name__ == "__main__":

    proses_kirim_serial("inisiasi")

    proses_logika()

            







